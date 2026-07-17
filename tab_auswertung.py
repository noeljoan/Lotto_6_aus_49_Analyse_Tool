"""
Tab 3: Auswertung – Treffer und Gewinne pro Ziehung
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Durchschnittliche Gewinnbeträge je Treffer (ohne Superzahl)
GEWINNE = {
    6: 500000.00,
    5: 4000.00,
    4: 50.00,
    3: 10.90,
}
EINSATZ = 1.20  # EUR pro Grille pro Ziehung


def berechne_treffer(grille, ziehung):
    return len(set(grille) & set(ziehung))


def analysiere(grilles, ziehungen):
    ergebnisse = []
    for z in ziehungen:
        zahlen = z["zahlen"]
        treffer = {3: 0, 4: 0, 5: 0, 6: 0}
        gewinn = 0.0
        for g in grilles:
            t = berechne_treffer(g, zahlen)
            if t >= 3:
                treffer[t] += 1
                gewinn += GEWINNE.get(t, 0)
        einsatz = len(grilles) * EINSATZ
        ergebnisse.append({
            "datum":   z["datum"],
            "zahlen":  zahlen,
            "einsatz": einsatz,
            "3er":     treffer[3],
            "4er":     treffer[4],
            "5er":     treffer[5],
            "6er":     treffer[6],
            "gewinn":  gewinn,
            "bilanz":  gewinn - einsatz,
        })
    return ergebnisse


class AuswertungTab(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg="#F5F5F5")
        self.app = app
        self._ergebnisse = []
        self._build_ui()

    def _build_ui(self):
        # --- Toolbar ---
        bar = tk.Frame(self, bg="#E8E8E8", pady=6)
        bar.pack(fill="x", padx=10, pady=(10, 0))

        tk.Button(bar, text="  Auswertung starten  ", command=self.refresh,
                  bg="#2E75B6", fg="white", font=("Segoe UI", 10, "bold"),
                  relief="flat", cursor="hand2", padx=8, pady=4).pack(side="left", padx=6)

        tk.Button(bar, text="  Excel exportieren  ", command=self._export_excel,
                  bg="#1F4E79", fg="white", font=("Segoe UI", 10, "bold"),
                  relief="flat", cursor="hand2", padx=8, pady=4).pack(side="left", padx=6)

        self.lbl_status = tk.Label(bar, text="", bg="#E8E8E8",
                                    fg="#1F4E79", font=("Segoe UI", 9, "italic"))
        self.lbl_status.pack(side="left", padx=16)

        # --- Zusammenfassung ---
        summ = tk.Frame(self, bg="#F5F5F5")
        summ.pack(fill="x", padx=10, pady=8)

        self.kpi_vars = {}
        kpis = [
            ("Ziehungen", "ziehungen"),
            ("Grilles", "grilles"),
            ("Gesamteinsatz", "einsatz"),
            ("Gesamtgewinn", "gewinn"),
            ("Netto-Bilanz", "bilanz"),
            ("3er Treffer", "t3"),
            ("4er Treffer", "t4"),
            ("5er Treffer", "t5"),
            ("6er Treffer", "t6"),
        ]
        for label, key in kpis:
            box = tk.Frame(summ, bg="white", relief="solid", bd=1, padx=12, pady=8)
            box.pack(side="left", padx=4, expand=True, fill="x")
            tk.Label(box, text=label, bg="white", fg="gray",
                     font=("Segoe UI", 8)).pack()
            var = tk.StringVar(value="–")
            self.kpi_vars[key] = var
            tk.Label(box, textvariable=var, bg="white", fg="#1F4E79",
                     font=("Segoe UI", 11, "bold")).pack()

        # --- Tabelle ---
        table_frame = tk.Frame(self, bg="#F5F5F5")
        table_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        cols = ("Datum", "Gezogene Zahlen", "Einsatz (EUR)",
                "3er", "4er", "5er", "6er", "Gewinn (EUR)", "Bilanz (EUR)")
        self.tree = ttk.Treeview(table_frame, columns=cols,
                                  show="headings", height=16)

        widths = [100, 260, 110, 60, 60, 60, 60, 110, 110]
        for col, w in zip(cols, widths):
            self.tree.heading(col, text=col,
                              command=lambda c=col: self._sort_by(c))
            self.tree.column(col, width=w, anchor="center")

        self.tree.tag_configure("gewinn", background="#E2EFDA")
        self.tree.tag_configure("kein", background="#FFFFFF")

        sb_y = ttk.Scrollbar(table_frame, orient="vertical",
                             command=self.tree.yview)
        sb_x = ttk.Scrollbar(table_frame, orient="horizontal",
                             command=self.tree.xview)
        self.tree.configure(yscrollcommand=sb_y.set,
                             xscrollcommand=sb_x.set)
        sb_y.pack(side="right", fill="y")
        sb_x.pack(side="bottom", fill="x")
        self.tree.pack(side="left", fill="both", expand=True)

    def refresh(self):
        if not self.app.grilles:
            self.lbl_status.config(text="Keine Grilles vorhanden!", fg="#C00000")
            return
        if not self.app.ziehungen:
            self.lbl_status.config(text="Keine Ziehungen vorhanden!", fg="#C00000")
            return

        self._ergebnisse = analysiere(self.app.grilles, self.app.ziehungen)
        self._populate_tree(self._ergebnisse)
        self._update_kpis(self._ergebnisse)
        self.lbl_status.config(
            text=f"Auswertung: {len(self._ergebnisse)} Ziehungen × "
                 f"{len(self.app.grilles)} Grilles", fg="#1F4E79")

    def _populate_tree(self, data):
        self.tree.delete(*self.tree.get_children())
        for r in data:
            nums = " – ".join(f"{n:02d}" for n in r["zahlen"])
            tag = "gewinn" if r["gewinn"] > 0 else "kein"
            self.tree.insert("", "end", values=(
                r["datum"], nums,
                f"{r['einsatz']:.2f}",
                r["3er"], r["4er"], r["5er"], r["6er"],
                f"{r['gewinn']:.2f}",
                f"{r['bilanz']:.2f}"
            ), tags=(tag,))

    def _update_kpis(self, data):
        total_einsatz = sum(r["einsatz"] for r in data)
        total_gewinn  = sum(r["gewinn"]  for r in data)
        bilanz        = total_gewinn - total_einsatz
        self.kpi_vars["ziehungen"].set(str(len(data)))
        self.kpi_vars["grilles"].set(str(len(self.app.grilles)))
        self.kpi_vars["einsatz"].set(f"{total_einsatz:,.2f} EUR")
        self.kpi_vars["gewinn"].set(f"{total_gewinn:,.2f} EUR")
        color = "#00B050" if bilanz >= 0 else "#C00000"
        self.kpi_vars["bilanz"].set(f"{bilanz:,.2f} EUR")
        self.kpi_vars["t3"].set(str(sum(r["3er"] for r in data)))
        self.kpi_vars["t4"].set(str(sum(r["4er"] for r in data)))
        self.kpi_vars["t5"].set(str(sum(r["5er"] for r in data)))
        self.kpi_vars["t6"].set(str(sum(r["6er"] for r in data)))

    def _sort_by(self, col):
        items = [(self.tree.set(k, col), k) for k in self.tree.get_children("")]
        try:
            items.sort(key=lambda x: float(x[0].replace(",", ".")))
        except ValueError:
            items.sort()
        for idx, (_, k) in enumerate(items):
            self.tree.move(k, "", idx)

    def _export_excel(self):
        if not self._ergebnisse:
            messagebox.showwarning("Hinweis", "Bitte zuerst die Auswertung starten.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel-Dateien", "*.xlsx")],
            initialfile="Lotto_Auswertung.xlsx"
        )
        if not path:
            return
        try:
            _export_to_excel(path, self._ergebnisse, self.app.grilles)
            messagebox.showinfo("Exportiert", f"Datei gespeichert:\n{path}")
        except Exception as e:
            messagebox.showerror("Fehler", f"Export fehlgeschlagen:\n{e}")


# ─── Excel Export ─────────────────────────────────────────────────────────────

def _export_to_excel(path, ergebnisse, grilles):
    wb = openpyxl.Workbook()
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    grn_fill  = PatternFill("solid", fgColor="E2EFDA")
    sum_fill  = PatternFill("solid", fgColor="BDD7EE")
    hdr_font  = Font(color="FFFFFF", bold=True, size=11)
    bold      = Font(bold=True)
    center    = Alignment(horizontal="center")
    EUR       = '#,##0.00 "EUR"'

    # --- Blatt 1: Auswertung ---
    ws = wb.active
    ws.title = "Auswertung"
    headers = ["Datum", "Gezogene Zahlen", "Einsatz (EUR)",
               "3er", "4er", "5er", "6er", "Gewinn (EUR)", "Bilanz (EUR)"]
    widths  = [14, 30, 14, 8, 8, 8, 8, 14, 14]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, c, h)
        cell.fill, cell.font = hdr_fill, hdr_font
        cell.alignment, cell.border = center, border
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 22

    for r_idx, r in enumerate(ergebnisse, 2):
        nums = ", ".join(map(str, r["zahlen"]))
        fill = grn_fill if r["gewinn"] > 0 else PatternFill()
        vals = [r["datum"], nums, r["einsatz"],
                r["3er"], r["4er"], r["5er"], r["6er"],
                r["gewinn"], r["bilanz"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r_idx, c, v)
            cell.fill, cell.alignment, cell.border = fill, center, border
            if c in (3, 8, 9):
                cell.number_format = EUR

    # Summenzeile
    sr = len(ergebnisse) + 2
    ws.cell(sr, 1, "GESAMT").font = bold
    for c, v in [
        (3,  sum(r["einsatz"] for r in ergebnisse)),
        (4,  sum(r["3er"]     for r in ergebnisse)),
        (5,  sum(r["4er"]     for r in ergebnisse)),
        (6,  sum(r["5er"]     for r in ergebnisse)),
        (7,  sum(r["6er"]     for r in ergebnisse)),
        (8,  sum(r["gewinn"]  for r in ergebnisse)),
        (9,  sum(r["bilanz"]  for r in ergebnisse)),
    ]:
        cell = ws.cell(sr, c, v)
        cell.fill, cell.font, cell.alignment, cell.border = sum_fill, bold, center, border
        if c in (3, 8, 9):
            cell.number_format = EUR

    # --- Blatt 2: Grilles ---
    ws2 = wb.create_sheet("Meine Grilles")
    for c, h in enumerate(["#", "Nummern"], 1):
        cell = ws2.cell(1, c, h)
        cell.fill, cell.font = hdr_fill, hdr_font
        cell.alignment = center
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 30
    for i, g in enumerate(grilles, 1):
        ws2.cell(i+1, 1, i).alignment = center
        ws2.cell(i+1, 2, ", ".join(f"{n:02d}" for n in g)).alignment = center

    wb.save(path)
